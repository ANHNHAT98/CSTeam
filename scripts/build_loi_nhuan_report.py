#!/usr/bin/env python3
"""
Xây file "Lợi nhuận Crs" (2 sheet: PAKD + Chi tiết theo tháng) từ file PAKD gốc do
người dùng upload + dữ liệu Timesheet đã tính toán ở trình duyệt.

Vì sao viết bằng Python/openpyxl thay vì làm hết ở trình duyệt (SheetJS)?
SheetJS bản miễn phí (đang dùng ở trang crs-loi-nhuan-du-an.html) chỉ ĐỌC được
style của cell (màu nền, màu chữ, in đậm...) nhưng KHÔNG GHI LẠI được khi xuất
file mới — mọi cell sẽ mất hết màu, chỉ giữ được định dạng số. openpyxl không có
giới hạn này nên việc build file cuối cùng được chuyển sang đây.

Cách gọi:
    python3 build_loi_nhuan_report.py <input.xlsx> <payload.json> <output.xlsx>

payload.json có dạng:
{
  "tiersUsed": ["Fresher","Junior","Middle","Senior"],
  "rateTable": {"Fresher":100000, "Junior":150000, ...},
  "monthRows": [{"month":"2025-07","cells":[80,0,0,16]}, ...]
}

In ra stdout 1 dòng JSON kết quả: {"ok":true, ...} hoặc {"ok":false,"error":"..."}
"""
import sys
import json
import re
import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def find_row_by_col_a_letter(ws, letter, max_row=80):
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == letter:
            return r
    return -1


def find_row_by_label_contains(ws, col_letter, needle_lower, max_row=80):
    col = ord(col_letter.upper()) - ord('A') + 1
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and needle_lower in str(v).lower():
            return r
    return -1


def resolve_profit_rows(ws):
    rows = {
        'revenue': find_row_by_label_contains(ws, 'B', 'doanh thu thuần'),
        'labor': find_row_by_label_contains(ws, 'B', 'nhân công'),
        'cost': find_row_by_col_a_letter(ws, 'C'),
        'profitBeforeTax': find_row_by_col_a_letter(ws, 'D'),
        'tax': find_row_by_col_a_letter(ws, 'E'),
        'profitAfterTax': find_row_by_col_a_letter(ws, 'F'),
        'profitPct': find_row_by_col_a_letter(ws, 'G'),
    }
    missing = [k for k, v in rows.items() if v == -1]
    if missing:
        raise ValueError(
            'Không tìm thấy dòng: ' + ', '.join(missing) +
            ' trong sheet PAKD — kiểm tra lại cấu trúc file (cột A phải có A/B/C/D/E/F/G, '
            'cột B phải có dòng "Doanh thu thuần" và dòng chứa "nhân công").'
        )
    return rows


def extract_sum_range(formula):
    if not isinstance(formula, str):
        return None
    m = re.search(r'SUM\(\$?[A-Z]\$?(\d+):\$?[A-Z]\$?(\d+)\)', formula, re.IGNORECASE)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def mirror_formula_to_actual(formula):
    """Copy công thức của cột D (Kế hoạch) sang cột E (Thực tế): thay mọi tham
    chiếu tới cột D bằng cột E, giữ nguyên số dòng và dấu $ nếu có."""
    if not isinstance(formula, str):
        return None
    return re.sub(r'(\$?)D(\$?)(\d+)', r'\1E\2\3', formula)


def flatten_cross_sheet_formulas(ws_formula, ws_value, max_row=80, max_col=10):
    """Sheet PAKD gốc có công thức tham chiếu sang sheet khác (Quotation...) mà
    file xuất không giữ lại (chỉ giữ PAKD + Chi tiết theo tháng). Đóng băng các
    công thức đó thành giá trị (lấy từ bản đọc data_only=True), nếu không sẽ ra
    #NAME?/#REF! vì sheet đích không còn tồn tại."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws_formula.cell(row=r, column=c)
            f = cell.value
            if isinstance(f, str) and f.startswith('=') and '!' in f:
                cached = ws_value.cell(row=r, column=c).value
                cell.value = cached if cached is not None else 0


def set_formula(ws, ref, formula, keep_style_from=None):
    cell = ws[ref]
    cell.value = '=' + formula if not formula.startswith('=') else formula
    if keep_style_from is not None:
        src = ws[keep_style_from]
        cell.font = copy.copy(src.font)
        cell.fill = copy.copy(src.fill)
        cell.number_format = src.number_format
        cell.border = copy.copy(src.border)
        cell.alignment = copy.copy(src.alignment)


def main():
    if len(sys.argv) != 4:
        print(json.dumps({'ok': False, 'error': 'Cần đúng 3 tham số: input.xlsx payload.json output.xlsx'}))
        sys.exit(1)

    input_path, payload_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]

    with open(payload_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)

    tiers_used = payload['tiersUsed']
    rate_table = payload['rateTable']
    month_rows = payload['monthRows']  # [{month, cells:[...]}]

    wb_f = load_workbook(input_path, data_only=False)
    wb_v = load_workbook(input_path, data_only=True)

    if 'PAKD' in wb_f.sheetnames:
        profit_sheet_name = 'PAKD'
    else:
        candidates = [n for n in wb_f.sheetnames if 'PAKD' in n.upper()]
        if not candidates:
            raise ValueError('Không tìm thấy sheet PAKD trong file.')
        profit_sheet_name = candidates[0]

    ws_f = wb_f[profit_sheet_name]
    ws_v = wb_v[profit_sheet_name]

    # Cột "Thực tế" (E) trong file gốc thường rất hẹp vì trước đây luôn trống —
    # giờ có số liệu thật (dạng kế toán, có thể âm) nên nới rộng bằng cột "Kế hoạch" (D)
    # để không bị hiển thị "###".
    d_width = ws_f.column_dimensions['D'].width
    e_width = ws_f.column_dimensions['E'].width
    if d_width and (not e_width or e_width < d_width):
        ws_f.column_dimensions['E'].width = d_width

    rows_map = resolve_profit_rows(ws_f)

    def get_val(col, row):
        v = ws_v[f'{col}{row}'].value
        return v if isinstance(v, (int, float)) else 0

    revenue = get_val('D', rows_map['revenue'])
    plan_labor = get_val('D', rows_map['labor'])

    cost_formula = ws_f[f'D{rows_map["cost"]}'].value
    sum_range = extract_sum_range(cost_formula) or (rows_map['labor'], rows_map['labor'])

    other_cost = 0
    for r in range(sum_range[0], sum_range[1] + 1):
        if r == rows_map['labor']:
            continue
        other_cost += get_val('D', r)

    # ---- Đóng băng công thức liên-sheet (vd Quotation!) thành giá trị, TRƯỚC khi mirror ----
    flatten_cross_sheet_formulas(ws_f, ws_v, max_row=60, max_col=8)

    # ---------- Sheet "Chi tiết theo tháng" ----------
    month_sheet_name = 'Chi tiết theo tháng'
    if month_sheet_name in wb_f.sheetnames:
        del wb_f[month_sheet_name]
    ws_m = wb_f.create_sheet(month_sheet_name)

    navy = 'FF16365C'
    header_font = Font(name='Times New Roman', bold=True, color='FFFFFFFF')
    header_fill = PatternFill('solid', fgColor=navy)
    thin = Side(style='thin', color='BFBFBF')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    header = ['Tháng'] + [f'{t} (giờ)' for t in tiers_used] + ['Chi phí tháng (VNĐ)']
    ws_m.append(header)
    for c in range(1, len(header) + 1):
        cell = ws_m.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left' if c == 1 else 'right', vertical='center')
        cell.border = border_all

    n_tiers = len(tiers_used)
    cost_col_letter = get_column_letter(n_tiers + 2)

    total_labor = 0
    for i, r in enumerate(month_rows):
        excel_row = i + 2
        row_vals = [r['month']] + list(r['cells']) + [None]
        ws_m.append(row_vals)
        parts = []
        for ti, t in enumerate(tiers_used):
            hour_col = get_column_letter(ti + 2)
            rate = rate_table.get(t, 0)
            parts.append(f'{hour_col}{excel_row}*{rate}')
        month_cost = sum(h * rate_table.get(t, 0) for h, t in zip(r['cells'], tiers_used))
        total_labor += month_cost
        cost_cell = ws_m.cell(row=excel_row, column=n_tiers + 2)
        cost_cell.value = '=' + '+'.join(parts)
        cost_cell.number_format = '#,##0" đ"'
        cost_cell.border = border_all
        for ti in range(n_tiers):
            c = ws_m.cell(row=excel_row, column=ti + 2)
            c.number_format = '0.##'
            c.border = border_all
            c.alignment = Alignment(horizontal='right')
        ws_m.cell(row=excel_row, column=1).border = border_all

    total_row = len(month_rows) + 2
    ws_m.cell(row=total_row, column=1, value='Tổng').font = Font(name='Times New Roman', bold=True)
    total_cost_cell = ws_m.cell(row=total_row, column=n_tiers + 2)
    total_cost_cell.value = f'=SUM({cost_col_letter}2:{cost_col_letter}{total_row - 1})'
    total_cost_cell.number_format = '#,##0" đ"'
    total_cost_cell.font = Font(name='Times New Roman', bold=True)
    for c in range(1, n_tiers + 2):
        ws_m.cell(row=total_row, column=c).border = border_all
    ws_m.cell(row=total_row, column=n_tiers + 2).border = border_all

    note_row = total_row + 2
    ws_m.cell(row=note_row, column=1,
              value='Đơn giá áp dụng (VNĐ/giờ, lấy từ sheet Monitor của file PAKD lúc xuất file này):').font = \
        Font(name='Times New Roman', italic=True)
    for i, t in enumerate(tiers_used):
        ws_m.cell(row=note_row + 1 + i, column=1, value=t)
        ws_m.cell(row=note_row + 1 + i, column=2, value=rate_table.get(t, 0))

    ws_m.column_dimensions['A'].width = 12
    for ti in range(n_tiers):
        ws_m.column_dimensions[get_column_letter(ti + 2)].width = 14
    ws_m.column_dimensions[cost_col_letter].width = 20
    ws_m.freeze_panes = 'A2'

    total_cost_ref_for_link = f"{cost_col_letter}{total_row}"

    # ---------- Ghi công thức cột E (Thực tế) vào sheet PAKD, giữ nguyên style gốc ----------
    r_rev = rows_map['revenue']
    set_formula(ws_f, f'E{r_rev}', f'D{r_rev}')

    for r in range(sum_range[0], sum_range[1] + 1):
        if r == rows_map['labor']:
            continue
        f = ws_f[f'D{r}'].value
        if isinstance(f, str) and f.startswith('='):
            set_formula(ws_f, f'E{r}', mirror_formula_to_actual(f))

    set_formula(ws_f, f'E{rows_map["labor"]}', f"'{month_sheet_name}'!{total_cost_ref_for_link}")

    for key in ('cost', 'profitBeforeTax', 'tax', 'profitAfterTax', 'profitPct'):
        r = rows_map[key]
        f = ws_f[f'D{r}'].value
        if isinstance(f, str) and f.startswith('='):
            set_formula(ws_f, f'E{r}', mirror_formula_to_actual(f))

    # Sheet PAKD trong file gốc có thể đang ở trạng thái ẩn (state=hidden) — vì đây
    # là sheet DUY NHẤT còn lại nên phải hiện lên, nếu không Excel sẽ mở file trắng.
    ws_f.sheet_state = 'visible'

    # ---------- Xoá mọi sheet khác, chỉ giữ PAKD + Chi tiết theo tháng ----------
    for name in list(wb_f.sheetnames):
        if name not in (profit_sheet_name, month_sheet_name):
            del wb_f[name]

    wb_f.active = wb_f.sheetnames.index(profit_sheet_name)
    try:
        wb_f.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    wb_f.save(output_path)

    print(json.dumps({
        'ok': True,
        'profitSheetName': profit_sheet_name,
        'monthSheetName': month_sheet_name,
        'revenue': revenue,
        'planLabor': plan_labor,
        'otherCost': other_cost,
        'totalLabor': total_labor,
        'rowsMap': rows_map,
        'sumRange': list(sum_range),
    }, ensure_ascii=False))


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(json.dumps({'ok': False, 'error': str(e)}, ensure_ascii=False))
        sys.exit(1)
